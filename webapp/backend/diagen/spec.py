"""Load rule sheets + data from the spec workbook, or nodes/edges from CSVs."""
from __future__ import annotations
import csv
from openpyxl import load_workbook
from .model import ShapeDef, LineDef, Node, Edge, Diagram
from .defaults import DEFAULT_SHAPES, DEFAULT_LINES


def _f(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _i(v, default=1):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _s(v):
    return str(v).strip() if v is not None else ''


def _rows(ws):
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row
            continue
        if row[0] is None or _s(row[0]) == '':
            continue
        yield dict(zip([_s(h) for h in header], row))


class Spec:
    def __init__(self, path=None, *, use_defaults=False):
        """`path` (str/Path/file-like, xlsx) is optional so a Spec can also be
        built purely from in-memory rows (see `Spec.blank()` +
        `add_shape_row`/`add_line_row`/`load_rows`, used by the webapp's
        CSV-paste, template, and outline-mode flows).

        `use_defaults=True` pre-seeds `self.shapes`/`self.lines` with the
        ~40-entity/~25-relation built-in palette (`diagen.defaults`) ported
        from the JS playground, so those flows work without the caller ever
        needing to define a Shape_Library/Line_Rules sheet. Off by default so
        the CLI and every existing xlsx-driven diagram keep behaving exactly
        as before — those workbooks are fully self-contained and don't need
        (or expect) an implicit fallback palette.

        Custom rows never override a name that's already present (matches
        the JS `mergeRules` semantics: first definition wins) — a workbook
        can freely ADD new entity/relation types alongside the defaults, but
        can't silently reshape one of the built-ins out from under it.
        """
        self.shapes: dict[str, ShapeDef] = dict(DEFAULT_SHAPES) if use_defaults else {}
        self.lines: dict[str, LineDef] = dict(DEFAULT_LINES) if use_defaults else {}
        self.diagrams: dict[str, Diagram] = {}
        if path is not None:
            self._load_workbook(path)
        self._link()

    @classmethod
    def blank(cls, use_defaults=True) -> 'Spec':
        """A Spec with no workbook loaded — starting point for CSV-paste,
        template-generated, and outline-mode diagrams."""
        return cls(None, use_defaults=use_defaults)

    def _load_workbook(self, path):
        wb = load_workbook(path, data_only=False)
        if 'Shape_Library' in wb.sheetnames:
            for r in _rows(wb['Shape_Library']):
                self.add_shape_row(r)
        if 'Line_Rules' in wb.sheetnames:
            for r in _rows(wb['Line_Rules']):
                self.add_line_row(r)
        if 'Diagrams' in wb.sheetnames:
            for r in _rows(wb['Diagrams']):
                d = Diagram(id=_s(r['diagram_id']), scenario=_s(r['scenario_id']),
                            title=_s(r['title']), direction=_s(r.get('direction')),
                            theme=_s(r.get('theme')))
                self.diagrams[d.id] = d
        if 'Nodes' in wb.sheetnames:
            for r in _rows(wb['Nodes']):
                self._add_node(r)
        if 'Edges' in wb.sheetnames:
            for r in _rows(wb['Edges']):
                self._add_edge(r)

    # -- rule ingestion (additive-only; see docstring above) ------------
    def add_shape_row(self, r) -> bool:
        t = _s(r.get('entity_type'))
        if not t or t in self.shapes:
            return False
        self.shapes[t] = ShapeDef(
            entity_type=t, family=_s(r.get('family')),
            shape=_s(r.get('shape')),
            default_w=_f(r.get('default_w'), 120), default_h=_f(r.get('default_h'), 50),
            min_w=_f(r.get('min_w'), 80), min_h=_f(r.get('min_h'), 36),
            fill=_s(r.get('fill_hex')) or '#FFFFFF', stroke=_s(r.get('stroke_hex')) or '#333333',
            auto=_s(r.get('auto_size')) or 'Y')
        return True

    def add_line_row(self, r) -> bool:
        t = _s(r.get('relation_type'))
        if not t or t in self.lines:
            return False
        self.lines[t] = LineDef(
            relation_type=t, family=_s(r.get('family')),
            style=_s(r.get('line_style')) or 'solid', width=_f(r.get('width_px'), 1),
            source_end=_s(r.get('source_end')) or 'none', target_end=_s(r.get('target_end')) or 'none',
            routing=_s(r.get('routing')) or 'orthogonal', label_pos=_s(r.get('label_position')),
            color=_s(r.get('color_hex')) or '#333333')
        return True

    # -- data ingestion ------------------------------------------------
    def _add_node(self, r):
        n = Node(
            id=_s(r['node_id']), diagram=_s(r.get('diagram_id')) or 'D1',
            parent=_s(r.get('parent_id')),
            type=_s(r['entity_type']), label=_s(r.get('label')) or _s(r['node_id']),
            rank=_i(r.get('rank_hint'), 0), order=_i(r.get('order_hint'), 0),
            w_override=_f(r.get('w_override'), 0) or None,
            h_override=_f(r.get('h_override'), 0) or None,
            fill_override=_s(r.get('fill_override')),
            stroke_override=_s(r.get('stroke_override')), meta=_s(r.get('metadata')))
        d = self.diagrams.setdefault(n.diagram, Diagram(n.diagram, '', n.diagram))
        d.nodes[n.id] = n

    def _add_edge(self, r):
        e = Edge(id=_s(r.get('edge_id')), diagram=_s(r.get('diagram_id')) or 'D1',
                 source=_s(r['source_id']), target=_s(r['target_id']),
                 relation=_s(r['relation_type']), label=_s(r.get('label')),
                 hint=_s(r.get('waypoint_hint')),
                 sport=_s(r.get('source_port')), tport=_s(r.get('target_port')))
        d = self.diagrams.setdefault(e.diagram, Diagram(e.diagram, '', e.diagram))
        d.edges.append(e)

    def load_csv(self, nodes_csv, edges_csv=None, diagram_id=None, title=None):
        """Replace workbook sample data with external CSV data (same headers)."""
        self.diagrams = {}
        with open(nodes_csv, newline='', encoding='utf-8-sig') as f:
            for r in csv.DictReader(f):
                if r.get('node_id'):
                    self._add_node(r)
        if edges_csv:
            with open(edges_csv, newline='', encoding='utf-8-sig') as f:
                for r in csv.DictReader(f):
                    if r.get('source_id'):
                        self._add_edge(r)
        if title:
            for d in self.diagrams.values():
                d.title = title
        self._link()

    def load_rows(self, node_rows, edge_rows=None, shape_rows=None, line_rows=None,
                  title=None, replace=True) -> dict:
        """Build (or replace) diagram data from plain dict rows — shared by the
        webapp's CSV-paste import, template generator, and outline mode.
        Mirrors the JS playground's tolerant `rowsToState`: a node row whose
        entity_type isn't in `self.shapes` is skipped and reported back
        rather than raising, and an edge row is only kept if both endpoints
        resolved to a real node. Returns {'unknown_types': [...]}."""
        for r in shape_rows or []:
            self.add_shape_row(r)
        for r in line_rows or []:
            self.add_line_row(r)
        if replace:
            self.diagrams = {}
        unknown = set()
        for r in node_rows:
            if not _s(r.get('node_id')):
                continue
            etype = _s(r.get('entity_type'))
            if etype not in self.shapes:
                unknown.add(etype or '(empty)')
                continue
            self._add_node(r)
        ids = {nid for d in self.diagrams.values() for nid in d.nodes}
        for r in edge_rows or []:
            if _s(r.get('source_id')) in ids and _s(r.get('target_id')) in ids:
                self._add_edge(r)
        if title:
            for d in self.diagrams.values():
                d.title = title
        self._link()
        return {'unknown_types': sorted(unknown)}

    def replace_diagram_data(self, diagram_id, node_rows, edge_rows=None,
                              shape_rows=None, line_rows=None) -> list:
        """Clear and rebuild ONLY `diagram_id`'s nodes/edges from raw rows —
        every other diagram in this Spec is untouched. Used by the webapp's
        'apply pasted CSV to the current diagram' action. Returns the sorted
        list of unknown entity_types skipped (empty if all resolved)."""
        for r in shape_rows or []:
            self.add_shape_row(r)
        for r in line_rows or []:
            self.add_line_row(r)
        d = self.diagrams.get(diagram_id)
        if d is None:
            d = Diagram(diagram_id, '', diagram_id)
            self.diagrams[diagram_id] = d
        d.nodes, d.edges = {}, []
        unknown = set()
        for r in node_rows:
            if not _s(r.get('node_id')):
                continue
            etype = _s(r.get('entity_type'))
            if etype not in self.shapes:
                unknown.add(etype or '(empty)')
                continue
            row = dict(r)
            row['diagram_id'] = diagram_id
            self._add_node(row)
        for r in edge_rows or []:
            if _s(r.get('source_id')) in d.nodes and _s(r.get('target_id')) in d.nodes:
                row = dict(r)
                row['diagram_id'] = diagram_id
                self._add_edge(row)
        self._link()
        return sorted(unknown)

    def _link(self):
        for d in self.diagrams.values():
            for n in d.nodes.values():
                n.children = []
            for n in d.nodes.values():
                if n.parent and n.parent in d.nodes:
                    d.nodes[n.parent].children.append(n)

    def shape_of(self, node: Node) -> ShapeDef:
        if node.type not in self.shapes:
            raise KeyError(f"entity_type '{node.type}' not in Shape_Library (node {node.id})")
        return self.shapes[node.type]

    def line_of(self, edge: Edge) -> LineDef:
        if edge.relation not in self.lines:
            raise KeyError(f"relation_type '{edge.relation}' not in Line_Rules (edge {edge.id})")
        return self.lines[edge.relation]
